import csv
import os
from datetime import datetime
import pandas as pd
import pdfkit
import jinja2

import openpyxl
from pandas import read_csv

import sendingemails
import appconstants

csv_name = "datasource/missing_data.csv"


def findingMissingValues(input_file, Missing_Info):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.get_sheet_by_name('Sheet1')
    print(sheet.max_row)
    print(sheet.max_column)
    for row in range(sheet.max_row):
        for column in range(sheet.max_column):
            if sheet.cell(row=row + 1, column=column + 1).value == None or sheet.cell(row=row + 1,
                                                                                      column=column + 1).value == " ":
                file1 = open(Missing_Info, "a")
                file1.write("this is row #: " + str(row) + ", this is column #: " + str(column) + "\n")
                file1.write(str(sheet.cell(row=1, column=1).value) + " has value : " + str(
                    sheet.cell(row=row + 1, column=1).value) + ", " + str(
                    sheet.cell(row=1, column=column + 1).value) + " has value : " + str(
                    sheet.cell(row=row + 1, column=column + 1).value) + "\n")

                file1.close()

                with open(csv_name, 'a', newline='') as file:
                    writer = csv.writer(file)
                    # writer.writerow(["Row_No", "Column_No", "Column Name","TestCase Name"])
                    writer.writerow([str(row), str(column), str(sheet.cell(row=1, column=column + 1).value),
                                     str(sheet.cell(row=row + 1, column=1).value),
                                     datetime.now().strftime("%d/%m/%Y %H:%M:%S")])

                print("this is row #: ", row, ", this is column #: ", column)
                print(sheet.cell(row=1, column=1).value, " has value : ", sheet.cell(row=row + 1, column=1).value,
                      sheet.cell(row=1, column=column + 1).value, " has value : ",
                      sheet.cell(row=row + 1, column=column + 1).value)
    if os.path.isfile(appconstants.Missing_Info):
        file1 = open(Missing_Info, "a")
        file1.write("end of values from input data file and timestamp is : " + datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S") + "\n")
        file1.write("\n")
        file1.close()

    if os.path.isfile(csv_name):
        df = read_csv(csv_name)

        df.columns = ["Row_#", "Column_#", "Column Name", "TestCase Name", "Timestamp"]
        # print(df.columns)
        df.to_csv(csv_name, index=False)

    # Converting into a HTML page
    if os.path.isfile(csv_name):
        df = read_csv(csv_name)
        df1 = df.values.tolist()
        df1.insert(0, df.columns.values)
        df = pd.DataFrame(df1)
        # html = df.to_html(classes='Missing Data from Meta Data Library', index=False, header=False)
        # # print(html)
        # html_file = open("datasource/missingReport.html", "w")
        # html_file.write(html)
        # html_file.close()

        html_string = '''<html>
        <head>
        <link href="css/style.css" type="text/css" rel="stylesheet">
        <link href="css/bootstrap.min.css" rel="stylesheet">
        <link href="/stylesheets/css_Sanford.css" rel="stylesheet" type="text/css" />
        <style type="text/css"> 
        </style>
        </head>
        <body style="background: rgb(0, 255, 255,0.2); ">
        <h1 style="color:DarkSlateGray;text-align:center;margin-left:auto;margin-right:auto;padding: 10px;">
        Missing Data from Meta Data Library
        </h1>
        <div class="dataframe Missing_Data" style="
          text-align:center;
		  margin-left:auto;
		  margin-right:auto;
		  padding: 10px;width:40%;
		  position: relative;
		  display: flex;
		  justify-content: center;
		  vertical-align:middle
		  align=center;
		  float:none;
		  font-family: Verdana, sans-serif;
		  
		  font-weight: 500;
          border-collapse:collapse;
          background-color: rgba(0, 128, 0, 0.2); " >
          {table}
        </div>
        </body>
        </html>'''

        # OUTPUT AN HTML FILE
        # df.style.set_table_attributes('style="align-self: center;"')
        html_file = open("datasource/missingReport.html", "w")
        html_file.write(html_string.format(table=df.to_html(classes='Missing_Data', index=False, header=False))
                        .replace('<table border="1"','<table border="1" style="align-self: center;"'))

        html_file.close()
        # with open('datasource/myhtml.html', 'w') as f:
        #     f.write(html_string.format(table=df.to_html(classes='mystyle', index=False, header=False)))


findingMissingValues(appconstants.input_data, appconstants.Missing_Info)
print(os.path.isfile(appconstants.Missing_Info))
# if os.path.isfile(appconstants.Missing_Info):
#     sendingemails.emailnotification.sendingEmail(appconstants.Missing_Info)


# dd/mm/YY H:M:S
print(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))



# Converting HTML to PDF
# html_string = '''<html>
#    <head>
#    <link href="css/style.css" type="text/css" rel="stylesheet">
#    <link href="css/bootstrap.min.css" rel="stylesheet">
#    <link href="/stylesheets/css_Sanford.css" rel="stylesheet" type="text/css" />
#    <style type="text/css">
#      body{{background: rgb(0, 255, 255,0.2);
#      }}
#      table, th, td {{font-weight: 500;
#      padding-top:30%;
#      margin-top: 10px;
#      font-size:10pt; border:1px solid black;
#       border-collapse:collapse; text-align:left;
#       margin-left:auto;margin-right:auto;
#       background-color: rgba(0, 128, 0, 0.2);
#
#
#       }}
#
#      th, td,h1 {{padding: 10px;font-family: Verdana, sans-serif;}}
#      .Missing_Data tr:first-child td {{
#       font-weight: 600;
#       color:Grey;
#                   }}
#
#
#
#
#    </style>
#    </head>
#    <body>
#    <h1 style="color:DarkSlateGray;text-align:center;margin-left:auto;margin-right:auto;padding: 10px;">
#    Missing Data from Meta Data Library
#    </h1>
#    <div style="font-weight: 500;
#      padding-top:30%;
#      margin-top: 10px;
#      font-size:10pt; border:1px solid black;
#       border-collapse:collapse; text-align:left;
#       margin-left:auto;margin-right:auto;
#       background-color: rgba(0, 128, 0, 0.2); ">
#      {table}
#    </div>
#    </body>
#    </html>'''









# pdfkit.from_file('datasource/missingReport.html', 'datasource/missingReport.pdf')
# with open('datasource/missingReport.html') as f:
#     pdfkit.from_file(f, 'out.pdf')
# import pdfkit
#
#
# path_wkthmltopdf = "C:\\Folder\\where\\wkhtmltopdf.exe"
# config = pdfkit.configuration(wkhtmltopdf = path_wkthmltopdf)
#
# pdfkit.from_url("http://google.com", "out.pdf", configuration=config)